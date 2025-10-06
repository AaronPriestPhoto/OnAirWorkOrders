#!/usr/bin/env python3
"""
Work Order Generator for OnAir Airline Manager

This script generates optimized work orders for idle planes by chaining together
the most profitable jobs based on each plane's priority (Pay, XP, or Balanced).

Features:
- Chains jobs sequentially from destination to departure
- Respects configurable time limits (default 24 hours with 0.5hr epsilon)
- Prioritizes planes by efficiency: Pay > XP > Balanced, then by tail number
- Ensures no job reuse across work orders
- Sorts jobs within work orders by time remaining to minimize penalties
- Supports two-pass optimization for better time utilization
"""

import argparse
import json
import os
import sqlite3
import sys
import shutil
import tempfile
import webbrowser
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Optional, Set, Tuple, Any
from math import sqrt
from pathlib import Path

# Ensure working directory is always the script's directory
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

# Also ensure project root is on sys.path for imports
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

# Set up BASE_DIR for consistent file path references
BASE_DIR = Path(__file__).resolve().parent

from onair.config import load_config
from onair import db as db_mod
from performance_optimizer import PerformanceOptimizer


# ----------------------------
# Safe File Operations
# ----------------------------
def safe_file_operation(operation_func, file_path: str, operation_name: str, max_retries: int = 10):
    """
    Safely perform file operations with user prompts when file is in use.
    
    Args:
        operation_func: Function to execute (should return True on success)
        file_path: Path to the file
        operation_name: Description of the operation for user prompts
        max_retries: Maximum number of retry attempts
    
    Returns:
        True if operation succeeded, False if user cancelled
    """
    for attempt in range(max_retries):
        try:
            return operation_func()
        except (PermissionError, OSError) as e:
            if attempt < max_retries - 1:
                print(f"\n⚠️  Cannot {operation_name} '{file_path}' - file may be open in another program.")
                print(f"   Error: {e}")
                print(f"\nPlease close the file and press Enter to retry, or type 'q' and Enter to quit:")
                
                user_input = input().strip().lower()
                if user_input == 'q':
                    print("Operation cancelled by user.")
                    return False
                print("Retrying...")
            else:
                print(f"\n❌ Failed to {operation_name} '{file_path}' after {max_retries} attempts.")
                print("Please close the file manually and run the script again.")
                return False
        except KeyboardInterrupt:
            print(f"\n\nOperation cancelled by user (Ctrl+C) during {operation_name}.")
            return False
        except Exception as e:
            print(f"❌ Unexpected error during {operation_name}: {e}")
            return False
    
    return False


def safe_remove_file(file_path: str) -> bool:
    """Safely remove a file with user prompts if it's in use."""
    if not os.path.exists(file_path):
        return True
    
    def remove_operation():
        os.remove(file_path)
        return True
    
    return safe_file_operation(remove_operation, file_path, "remove file")


def open_file(file_path: str) -> bool:
    """Open file with the default application."""
    try:
        # Convert to absolute path for better compatibility
        abs_path = os.path.abspath(file_path)
        
        # Use webbrowser to open files (works cross-platform)
        webbrowser.open(f"file://{abs_path}")
        return True
    except Exception as e:
        print(f"Warning: Could not auto-open file: {e}")
        return False


@dataclass
class JobInfo:
    """Information about a job for work order planning."""
    job_id: str
    departure: str
    destination: str
    pay: float
    xp: float
    distance_nm: float
    pay_per_hour: float
    xp_per_hour: float
    balance_score: float
    flight_hours: float
    time_remaining_hours: float
    penalty_amount: float = 0.0
    adjusted_pay_per_hour: float = 0.0
    adjusted_pay_total: float = 0.0
    source: str = ""
    speed_kts: float = 0.0
    min_airport_size: Optional[int] = None
    route: str = ""
    legs_count: int = 0
    hub_penalty_hours: float = 0.0
    legs: List[Dict[str, Any]] = field(default_factory=list)
    # Multi-job support
    total_payload_lbs: float = 0.0  # Total cargo weight for this job
    delivery_waypoint: Optional[str] = None  # Which waypoint this job gets delivered at


@dataclass
class MultiJobLeg:
    """A leg that can carry multiple jobs simultaneously."""
    from_icao: str
    to_icao: str
    distance_nm: float
    jobs: List[JobInfo] = field(default_factory=list)
    total_payload_lbs: float = 0.0
    flight_hours: float = 0.0
    speed_kts: float = 0.0
    
    def add_job(self, job: JobInfo) -> None:
        """Add a job to this leg."""
        self.jobs.append(job)
        self.total_payload_lbs += job.total_payload_lbs
    
    def can_add_job(self, job: JobInfo, plane_spec: Dict[str, Any]) -> Tuple[bool, str]:
        """Check if a job can be added to this leg considering payload and range constraints."""
        # Check if job's route includes this leg
        if not self._job_uses_this_leg(job):
            return False, "Job doesn't use this leg"
        
        # Calculate new total payload
        new_total_payload = self.total_payload_lbs + job.total_payload_lbs
        
        # Check payload capacity for this distance
        max_payload = self._calculate_payload_range_limit(plane_spec, self.distance_nm)
        if new_total_payload > max_payload:
            return False, f"Combined payload {new_total_payload:.0f}lbs exceeds capacity {max_payload:.0f}lbs for {self.distance_nm:.0f}nm"
        
        return True, "OK"
    
    def _job_uses_this_leg(self, job: JobInfo) -> bool:
        """Check if a job's route includes this leg."""
        for leg in job.legs:
            if (leg['from_icao'].upper() == self.from_icao.upper() and 
                leg['to_icao'].upper() == self.to_icao.upper()):
                return True
        return False
    
    def _calculate_payload_range_limit(self, plane_spec: Dict[str, Any], distance_nm: float) -> float:
        """Calculate maximum payload for given distance (copied from score_jobs.py)."""
        max_range = max(plane_spec["range1_nm"] or 0, plane_spec["range2_nm"] or 0)
        max_payload = max(plane_spec["payload1_lbs"] or 0, plane_spec["payload2_lbs"] or 0)
        
        if distance_nm > max_range:
            return 0
        
        if plane_spec["range1_nm"] == plane_spec["range2_nm"]:
            return max_payload
        
        if plane_spec["payload1_lbs"] == plane_spec["payload2_lbs"]:
            return plane_spec["payload1_lbs"]
        
        # Sort the range/payload pairs
        if plane_spec["range1_nm"] > plane_spec["range2_nm"]:
            r1, p1 = plane_spec["range2_nm"], plane_spec["payload2_lbs"]
            r2, p2 = plane_spec["range1_nm"], plane_spec["payload1_lbs"]
        else:
            r1, p1 = plane_spec["range1_nm"], plane_spec["payload1_lbs"]
            r2, p2 = plane_spec["range2_nm"], plane_spec["payload2_lbs"]
        
        if r1 == r2:
            return max_payload if distance_nm <= r1 else 0
        
        if distance_nm <= r1:
            return max(p1, p2)
        elif distance_nm >= r2:
            return min(p1, p2)
        else:
            m = (p2 - p1) / (r2 - r1)
            calculated_payload = p1 + m * (distance_nm - r1)
            return min(calculated_payload, max_payload)


@dataclass
class PlaneInfo:
    """Information about a plane for work order planning."""
    plane_id: str
    registration: str
    plane_type: str
    priority: str  # 'pay', 'xp', or 'balance'
    current_location: str
    status: int
    feasible_jobs: List[JobInfo] = field(default_factory=list)


@dataclass
class WorkOrder:
    """A work order containing a sequence of chained jobs for a plane."""
    plane_id: str
    plane_registration: str
    plane_type: str
    priority: str
    jobs: List[JobInfo] = field(default_factory=list)
    multi_job_legs: List[MultiJobLeg] = field(default_factory=list)  # New: legs that can carry multiple jobs
    execution_order: List[Tuple[str, Any]] = field(default_factory=list)  # Track execution order: ("job", job) or ("multi_leg", leg)
    total_hours: float = 0.0
    total_pay: float = 0.0
    total_xp: float = 0.0
    total_adjusted_pay: float = 0.0
    
    def add_job(self, job: JobInfo, accumulated_hours: float = 0.0) -> None:
        """Add a job to the work order and update totals."""
        # Calculate penalty based on accumulated time
        job.penalty_amount = self._calculate_penalty(job, accumulated_hours + job.flight_hours)
        job.adjusted_pay_total = job.pay - job.penalty_amount
        job.adjusted_pay_per_hour = job.adjusted_pay_total / job.flight_hours if job.flight_hours > 0 else 0.0
        
        self.jobs.append(job)
        self.execution_order.append(("job", job))  # Track execution order
        self.total_hours += job.flight_hours
        self.total_pay += job.pay
        self.total_xp += job.xp
        self.total_adjusted_pay += job.adjusted_pay_total
    
    def _calculate_penalty(self, job: JobInfo, hours_from_now: float, penalty_hours_to_max: float = 24.0) -> float:
        """Calculate penalty for a job based on when it will be completed."""
        hours_late = max(0.0, hours_from_now - job.time_remaining_hours)
        if hours_late > 0:
            penalty_percentage = min(1.0, hours_late / penalty_hours_to_max)
            # We need to get the max penalty from the job data - for now use a reasonable estimate
            max_penalty = job.pay * 0.5  # Assume max penalty is 50% of pay
            return max_penalty * penalty_percentage
        return 0.0
    
    def get_current_location(self) -> str:
        """Get the current location after all jobs in the work order."""
        # Check multi-job legs first (they come after regular jobs)
        if self.multi_job_legs:
            return self.multi_job_legs[-1].to_icao
        elif self.jobs:
            return self.jobs[-1].destination
        return ""
    
    def can_add_job(self, job: JobInfo, max_hours: float, epsilon_hours: float = 0.5) -> bool:
        """Check if a job can be added within time constraints."""
        new_total = self.total_hours + job.flight_hours
        return new_total <= (max_hours + epsilon_hours)
    
    def get_priority_score(self, job: JobInfo) -> float:
        """Get the priority score for a job based on plane priority."""
        if self.priority == "pay":
            # Use adjusted pay per hour if it's been calculated, otherwise use regular pay per hour
            return job.adjusted_pay_per_hour if job.adjusted_pay_per_hour > 0 else job.pay_per_hour
        elif self.priority == "xp":
            return job.xp_per_hour
        else:  # balance
            return job.balance_score
    
    def add_multi_job_leg(self, leg: MultiJobLeg) -> None:
        """Add a multi-job leg to the work order."""
        self.multi_job_legs.append(leg)
        self.execution_order.append(("multi_leg", leg))  # Track execution order
        
        # Recalculate job metrics for multi-job legs
        self._recalculate_multi_job_leg_metrics(leg)
        
        # Update totals based on all jobs in this leg
        for job in leg.jobs:
            self.total_pay += job.pay
            self.total_xp += job.xp
            self.total_adjusted_pay += job.adjusted_pay_total
        self.total_hours += leg.flight_hours
    
    def _recalculate_multi_job_leg_metrics(self, leg: MultiJobLeg) -> None:
        """Recalculate pay/XP per hour metrics for jobs in a multi-job leg."""
        if leg.flight_hours <= 0:
            return
        
        # Calculate combined totals for the leg
        total_pay = sum(job.pay for job in leg.jobs)
        total_xp = sum(job.xp for job in leg.jobs)
        
        # Calculate combined metrics
        combined_pay_per_hour = total_pay / leg.flight_hours
        combined_xp_per_hour = total_xp / leg.flight_hours
        combined_balance_score = (combined_pay_per_hour / 1_000_000 + combined_xp_per_hour / 100) / 2 if (combined_pay_per_hour > 0 and combined_xp_per_hour > 0) else (combined_pay_per_hour or combined_xp_per_hour)
        
        # Update each job's metrics to reflect the combined leg performance
        for job in leg.jobs:
            # Calculate this job's share of the combined metrics
            job_share = job.pay / total_pay if total_pay > 0 else 1.0 / len(leg.jobs)
            
            # Update the job's per-hour metrics to reflect the actual leg performance
            job.pay_per_hour = combined_pay_per_hour * job_share
            job.xp_per_hour = combined_xp_per_hour * job_share
            job.balance_score = combined_balance_score * job_share
    
    def can_add_job_to_leg(self, job: JobInfo, leg: MultiJobLeg, plane_spec: Dict[str, Any]) -> Tuple[bool, str]:
        """Check if a job can be added to an existing leg."""
        return leg.can_add_job(job, plane_spec)
    
    def get_all_jobs(self) -> List[JobInfo]:
        """Get all jobs from both regular jobs and multi-job legs."""
        all_jobs = list(self.jobs)
        for leg in self.multi_job_legs:
            all_jobs.extend(leg.jobs)
        return all_jobs
    
    def get_all_job_ids(self) -> Set[str]:
        """Get all job IDs from both regular jobs and multi-job legs."""
        job_ids = set()
        for job in self.jobs:
            job_ids.add(job.job_id)
        for leg in self.multi_job_legs:
            for job in leg.jobs:
                job_ids.add(job.job_id)
        return job_ids
    
    def get_final_destination(self) -> str:
        """Get the final destination of the work order."""
        if self.multi_job_legs:
            return self.multi_job_legs[-1].to_icao
        elif self.jobs:
            return self.jobs[-1].destination
        return ""


class WorkOrderGenerator:
    """Generates optimized work orders for idle planes."""
    
    def __init__(self, db_path: str, optimizer: Optional[PerformanceOptimizer] = None):
        self.original_db_path = db_path
        self.optimizer = optimizer
        self.used_jobs: Set[str] = set()
        
        # Performance optimization: Load entire database into RAM
        self._ram_db_path = self._create_ram_database()
        self._used_jobs_tracking: Set[str] = set()  # Track used jobs in memory only
        
        # Cache for expensive calculations
        self._distance_cache: Dict[Tuple[str, str], float] = {}
        self._airport_size_cache: Dict[str, int] = {}
        
        print(f"Database loaded into RAM at: {self._ram_db_path}")
    
    @property
    def db_path(self) -> str:
        """Get the current database path (RAM database during processing)."""
        return self._ram_db_path
    
    def _create_ram_database(self) -> str:
        """Create a copy of the database in RAM for ultra-fast access."""
        print("Creating RAM database for ultra-fast processing...")
        
        # Use a more efficient approach: create in temp directory but avoid unnecessary copying
        temp_fd, temp_path = tempfile.mkstemp(suffix='.db', prefix='onair_ram_', dir=tempfile.gettempdir())
        os.close(temp_fd)  # Close the file descriptor, we'll use the path
        
        try:
            # Copy the original database to the temporary location
            # Use copyfile instead of copy2 to avoid copying metadata
            shutil.copyfile(self.original_db_path, temp_path)
            print(f"Database copied to RAM: {temp_path}")
            return temp_path
        except Exception as e:
            print(f"Error creating RAM database: {e}")
            # Fallback to original database
            return self.original_db_path
    
    def _save_ram_database_to_disk(self) -> None:
        """Save the RAM database back to the original disk location."""
        if self._ram_db_path == self.original_db_path:
            return  # No need to save if we're using the original database
        
        print("Saving RAM database back to disk...")
        try:
            # Only create backup if the original database exists and is different
            if os.path.exists(self.original_db_path):
                backup_path = f"{self.original_db_path}.backup"
                shutil.copyfile(self.original_db_path, backup_path)
                print(f"Created backup: {backup_path}")
            
            # Copy the RAM database back to the original location
            shutil.copyfile(self._ram_db_path, self.original_db_path)
            print(f"Saved RAM database to: {self.original_db_path}")
            
            # Clean up the temporary RAM database
            os.unlink(self._ram_db_path)
            print("Cleaned up RAM database")
            
        except Exception as e:
            print(f"Error saving RAM database: {e}")
            print("Original database backup should be available")
    
    
    def write_used_jobs_to_database(self) -> None:
        """Write used jobs to database and save entire RAM database back to disk."""
        if self._used_jobs_tracking:
            print(f"Writing {len(self._used_jobs_tracking)} used jobs to database...")
            with db_mod.connect(self.db_path) as conn:
                # Batch update job_scores table to mark jobs as used
                job_ids = list(self._used_jobs_tracking)
                conn.executemany("""
                    UPDATE job_scores 
                    SET feasible = 0 
                    WHERE job_id = ?
                """, [(job_id,) for job_id in job_ids])
                conn.commit()
            print(f"Successfully marked {len(self._used_jobs_tracking)} jobs as used in database")
        
        # Save the entire RAM database back to disk
        self._save_ram_database_to_disk()
        
    def get_idle_planes(self) -> List[PlaneInfo]:
        """Get all idle planes sorted by priority and registration."""
        with db_mod.connect(self.db_path) as conn:
            # Get planes with status 0 (available/idle)
            rows = conn.execute("""
                SELECT id, registration, type, data_json 
                FROM airplanes 
                ORDER BY registration
            """).fetchall()
            
            planes = []
            for plane_id, registration, plane_type, data_json in rows:
                try:
                    data = json.loads(data_json) if data_json else {}
                    aircraft_status = data.get('AircraftStatus', 0)
                    
                    # Only include available/idle planes (status 0)
                    if aircraft_status != 0:
                        continue
                    
                    # Get current location
                    current_location = data.get('CurrentAirportICAO', '')
                    if not current_location:
                        # Try to find nearest airport using coordinates
                        lat = data.get('Latitude')
                        lon = data.get('Longitude')
                        if lat is not None and lon is not None:
                            current_location = self._find_nearest_airport(conn, float(lat), float(lon))
                        
                        if not current_location:
                            continue  # Skip planes without location
                    
                    # Get plane priority from specs
                    priority = self._get_plane_priority(conn, plane_type)
                    
                    plane = PlaneInfo(
                        plane_id=plane_id,
                        registration=registration or plane_id,
                        plane_type=plane_type or 'Unknown',
                        priority=priority,
                        current_location=current_location,
                        status=aircraft_status
                    )
                    
                    planes.append(plane)
                    
                except Exception as e:
                    print(f"Warning: Error processing plane {plane_id}: {e}")
                    continue
            
            # Sort by priority (Pay > XP > Balanced) then by registration
            priority_order = {'pay': 0, 'xp': 1, 'balance': 2}
            planes.sort(key=lambda p: (priority_order.get(p.priority, 3), p.registration))
            
            return planes
    
    def _get_plane_priority(self, conn: sqlite3.Connection, plane_type: str) -> str:
        """Get the priority setting for a plane type."""
        row = conn.execute("SELECT priority FROM plane_specs WHERE plane_type = ?", (plane_type,)).fetchone()
        if row and row[0]:
            priority = str(row[0]).strip().lower()
            if priority in ("pay", "xp", "balance"):
                return priority
        return "balance"  # Default
    
    def _find_nearest_airport(self, conn: sqlite3.Connection, lat: float, lon: float) -> Optional[str]:
        """Find the nearest airport to given coordinates."""
        # Get all airports with coordinates
        airports = conn.execute("""
            SELECT icao, latitude, longitude 
            FROM airports 
            WHERE latitude IS NOT NULL AND longitude IS NOT NULL
        """).fetchall()
        
        if not airports:
            return None
        
        min_distance = float('inf')
        nearest_icao = None
        
        for icao, airport_lat, airport_lon in airports:
            # Simple distance calculation (good enough for finding nearest)
            distance = sqrt((lat - airport_lat)**2 + (lon - airport_lon)**2)
            if distance < min_distance:
                min_distance = distance
                nearest_icao = icao
        
        return nearest_icao
    
    def _calculate_distance_between_airports(self, icao1: str, icao2: str) -> float:
        """Calculate distance between two airports using cached Haversine formula."""
        # Check cache first
        cache_key = (icao1.upper(), icao2.upper())
        if cache_key in self._distance_cache:
            return self._distance_cache[cache_key]
        
        # Direct database query (now using RAM database for speed)
        with db_mod.connect(self.db_path) as conn:
            # Get coordinates for both airports
            coords1 = conn.execute("SELECT latitude, longitude FROM airports WHERE icao = ?", (icao1,)).fetchone()
            coords2 = conn.execute("SELECT latitude, longitude FROM airports WHERE icao = ?", (icao2,)).fetchone()
            
            if not coords1 or not coords2:
                self._distance_cache[cache_key] = 0.0
                return 0.0
            
            lat1, lon1 = coords1
            lat2, lon2 = coords2
        
        # Haversine formula for great circle distance
        from math import radians, cos, sin, asin, sqrt
        
        # Convert to radians
        lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
        
        # Haversine formula
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
        c = 2 * asin(sqrt(a))
        
        # Earth's radius in nautical miles
        r = 3440.065  # nm
        distance = c * r
        
        # Cache the result
        self._distance_cache[cache_key] = distance
        return distance
    
    def load_feasible_jobs_for_plane(self, plane: PlaneInfo) -> None:
        """Load all feasible jobs for a specific plane."""
        with db_mod.connect(self.db_path) as conn:
            # Get feasible jobs for this plane
            rows = conn.execute("""
                SELECT js.job_id, js.pay_per_hour, js.xp_per_hour, js.balance_score,
                       j.departure, j.destination, j.pay, j.xp, j.computed_distance_nm, j.source
                FROM job_scores js
                JOIN jobs j ON j.id = js.job_id
                WHERE js.plane_id = ? AND js.feasible = 1
                ORDER BY js.balance_score DESC
            """, (plane.plane_id,)).fetchall()
            
            jobs = []
            for row in rows:
                job_id, pph, xph, balance, dep, dest, pay, xp, dist, source = row
                
                # Skip jobs that are already used (using memory tracking)
                if job_id in self._used_jobs_tracking:
                    continue
                
                # Get job legs
                legs = self._get_job_legs(conn, job_id)
                
                # Calculate flight hours and speed
                flight_hours, avg_speed = self._calculate_flight_hours_and_speed(plane, legs)
                
                # Calculate time remaining
                time_remaining = self._calculate_time_remaining_hours(conn, job_id)
                
                # Get actual departure and destination from legs
                actual_departure = legs[0]['from_icao'] if legs else (dep or '')
                actual_destination = legs[-1]['to_icao'] if legs else (dest or '')
                
                # Calculate route and min airport size
                route_str, min_airport_size = self._calculate_route_info(conn, legs)
                
                # Calculate total payload for this job
                total_payload = sum(leg.get('cargo_lbs', 0.0) for leg in legs)
                
                job_info = JobInfo(
                    job_id=job_id,
                    departure=actual_departure,
                    destination=actual_destination,
                    pay=pay or 0.0,
                    xp=xp or 0.0,
                    distance_nm=dist or 0.0,
                    pay_per_hour=pph or 0.0,
                    xp_per_hour=xph or 0.0,
                    balance_score=balance or 0.0,
                    flight_hours=flight_hours,
                    time_remaining_hours=time_remaining,
                    source=source or '',
                    speed_kts=avg_speed,
                    min_airport_size=min_airport_size,
                    route=route_str,
                    legs_count=len(legs),
                    legs=legs,
                    total_payload_lbs=total_payload
                )
                
                # Calculate adjusted pay values (penalty_amount defaults to 0.0)
                job_info.adjusted_pay_total = job_info.pay - job_info.penalty_amount
                job_info.adjusted_pay_per_hour = job_info.adjusted_pay_total / job_info.flight_hours if job_info.flight_hours > 0 else 0.0
                
                jobs.append(job_info)
            
            plane.feasible_jobs = jobs
    
    def _get_job_legs(self, conn: sqlite3.Connection, job_id: str) -> List[Dict[str, Any]]:
        """Get legs for a job."""
        rows = conn.execute("""
            SELECT leg_index, from_icao, to_icao, distance_nm, cargo_lbs
            FROM job_legs 
            WHERE job_id = ? 
            ORDER BY leg_index
        """, (job_id,)).fetchall()
        
        return [
            {
                'leg_index': row[0],
                'from_icao': row[1] or '',
                'to_icao': row[2] or '',
                'distance_nm': row[3] or 0.0,
                'cargo_lbs': row[4] or 0.0
            }
            for row in rows
        ]
    
    def _calculate_flight_hours_and_speed(self, plane: PlaneInfo, legs: List[Dict[str, Any]]) -> Tuple[float, float]:
        """Calculate total flight hours and average speed for job legs."""
        if not legs:
            return 0.0, 0.0
        
        total_hours = 0.0
        total_distance = 0.0
        
        for leg in legs:
            distance = leg['distance_nm']
            payload = leg['cargo_lbs']
            total_distance += distance
            
            # Get optimized speed if available
            if self.optimizer and plane.plane_type:
                # For airport sizes, we'll use a default or try to look them up
                dep_size = 3  # Default medium airport
                dest_size = 3  # Default medium airport
                
                speed = self.optimizer.get_optimized_speed(
                    plane.plane_type, distance, payload, dep_size, dest_size
                )
                
                # If optimizer returns 0 (plane not in performance curves), fall back to base speed
                if speed <= 0:
                    speed = self._get_base_speed(plane.plane_type)
            else:
                # Fallback to base speed from plane specs
                speed = self._get_base_speed(plane.plane_type)
            
            # Final safety check - ensure we have a valid speed
            if speed <= 0:
                print(f"Warning: Invalid speed {speed} for plane {plane.plane_type}, using default fallback")
                speed = 200.0  # Absolute fallback speed
            
            # Calculate flight hours for this leg
            if speed > 0 and distance > 0:
                leg_hours = distance / speed
                total_hours += leg_hours
            elif distance > 0:
                # If we have distance but no speed, something is wrong
                print(f"Warning: Distance {distance}nm but speed {speed}kts for plane {plane.plane_type}")
        
        # Calculate average speed
        if total_hours > 0:
            avg_speed = total_distance / total_hours
        else:
            # If no hours calculated, use base speed as fallback
            avg_speed = self._get_base_speed(plane.plane_type)
        
        return total_hours, avg_speed
    
    def _calculate_flight_hours(self, plane: PlaneInfo, legs: List[Dict[str, Any]]) -> float:
        """Calculate total flight hours for job legs (backward compatibility)."""
        flight_hours, _ = self._calculate_flight_hours_and_speed(plane, legs)
        return flight_hours
    
    def _calculate_route_info(self, conn: sqlite3.Connection, legs: List[Dict[str, Any]]) -> Tuple[str, Optional[int]]:
        """Calculate route string and minimum airport size from legs."""
        if not legs:
            return "", None
        
        # Build route string
        route_points = []
        min_airport_size = None
        
        for i, leg in enumerate(legs):
            from_icao = leg['from_icao']
            to_icao = leg['to_icao']
            
            # Add departure airport for first leg
            if i == 0 and from_icao:
                route_points.append(from_icao)
            
            # Add destination airport
            if to_icao:
                route_points.append(to_icao)
            
            # Check airport sizes
            for icao in [from_icao, to_icao]:
                if icao:
                    row = conn.execute("SELECT size FROM airports WHERE icao = ?", (icao,)).fetchone()
                    if row and row[0] is not None:
                        size = int(row[0])
                        min_airport_size = size if (min_airport_size is None or size < min_airport_size) else min_airport_size
        
        # Remove consecutive duplicates
        unique_route = []
        for point in route_points:
            if not unique_route or unique_route[-1] != point:
                unique_route.append(point)
        
        route_str = " -> ".join(unique_route)
        return route_str, min_airport_size
    
    def _get_base_speed(self, plane_type: str) -> float:
        """Get base cruise speed for a plane type."""
        # Direct database query (now using RAM database for speed)
        with db_mod.connect(self.db_path) as conn:
            # First try plane_specs table (optimization data from planes.xlsx)
            row = conn.execute("SELECT cruise_speed_kts FROM plane_specs WHERE plane_type = ?", (plane_type,)).fetchone()
            if row and row[0]:
                return float(row[0])
            
            # Fallback to airplanes table (fetched API data)
            row = conn.execute("SELECT data_json FROM airplanes WHERE type = ? LIMIT 1", (plane_type,)).fetchone()
            if row and row[0]:
                try:
                    data = json.loads(row[0])
                    # Try AircraftType.designSpeedVC first (most accurate)
                    aircraft_type = data.get('AircraftType', {})
                    speed = aircraft_type.get('designSpeedVC')
                    if speed and speed > 0:
                        return float(speed)
                    
                    # Fallback to direct CruiseSpeed field if available
                    speed = data.get('CruiseSpeed')
                    if speed and speed > 0:
                        return float(speed)
                except (json.JSONDecodeError, ValueError, TypeError):
                    pass
        
        return 200.0  # Default fallback speed
    
    def _calculate_time_remaining_hours(self, conn: sqlite3.Connection, job_id: str) -> float:
        """Calculate time remaining until job expiration in hours."""
        row = conn.execute("SELECT data_json FROM jobs WHERE id = ?", (job_id,)).fetchone()
        if not row or not row[0]:
            return 0.0
        
        try:
            job_data = json.loads(row[0])
            expiration_str = job_data.get("ExpirationDate")
            if not expiration_str:
                return 0.0
            
            # Parse the expiration date
            expiration_dt = datetime.fromisoformat(expiration_str.replace('Z', '+00:00'))
            if expiration_dt.tzinfo is None:
                expiration_dt = expiration_dt.replace(tzinfo=timezone.utc)
            
            # Calculate time remaining
            now = datetime.now(timezone.utc)
            time_remaining = expiration_dt - now
            
            return time_remaining.total_seconds() / 3600.0
        except Exception:
            return 0.0
    
    def find_chainable_jobs(self, plane: PlaneInfo, current_location: str, 
                           remaining_hours: float, epsilon_hours: float = 0.5) -> List[JobInfo]:
        """Find jobs that can be chained from the current location."""
        chainable = []
        
        # Limit to top 100 jobs for better quality while maintaining performance
        available_jobs = [job for job in plane.feasible_jobs if job.job_id not in self._used_jobs_tracking]
        available_jobs.sort(key=lambda j: j.balance_score, reverse=True)
        limited_jobs = available_jobs[:100]  # Consider top 100 jobs for better quality
        
        for job in limited_jobs:
            # Check if job departs from current location (use first leg)
            if not job.legs or job.legs[0]['from_icao'].upper() != current_location.upper():
                continue
            
            # Check if job fits within remaining time (with epsilon)
            if job.flight_hours > (remaining_hours + epsilon_hours):
                continue
            
            chainable.append(job)
        
        return chainable
    
    def find_jobs_from_major_hubs(self, plane: PlaneInfo, max_hours: float = 24.0) -> List[JobInfo]:
        """Find jobs from major job hubs that the plane could fly to first."""
        # Direct database query (now using RAM database for speed)
        with db_mod.connect(self.db_path) as conn:
            top_hubs = conn.execute("""
                SELECT from_icao, COUNT(DISTINCT job_id) as job_count 
                FROM job_legs 
                WHERE from_icao IS NOT NULL 
                GROUP BY from_icao 
                ORDER BY job_count DESC 
                LIMIT 50
            """).fetchall()
        
        hub_jobs = []
        
        # Debug for BONE planes
        if 'BONE' in plane.registration:
            print(f"    DEBUG {plane.registration}: Checking {len(top_hubs)} top hubs")
            for i, (hub_icao, job_count) in enumerate(top_hubs[:10]):  # Show top 10 hubs
                if hub_icao == plane.current_location:
                    continue
                hub_jobs_for_this_hub = []
                for job in plane.feasible_jobs:
                    if job.job_id in self._used_jobs_tracking:
                        continue
                    if not job.legs or job.legs[0]['from_icao'].upper() != hub_icao.upper():
                        continue
                    hub_jobs_for_this_hub.append(job)
                print(f"    DEBUG {plane.registration}: Hub {hub_icao} has {len(hub_jobs_for_this_hub)} available jobs")
        
        for hub_icao, job_count in top_hubs:
            if hub_icao == plane.current_location:
                continue  # Skip if already at this hub
            
            # Find jobs from this hub that are feasible for this plane
            for job in plane.feasible_jobs:
                if job.job_id in self._used_jobs_tracking:
                    continue
                
                if not job.legs or job.legs[0]['from_icao'].upper() != hub_icao.upper():
                    continue
                
                # Add a small penalty for having to fly to the hub first
                # This will be factored into the job selection
                job.hub_penalty_hours = self._calculate_hub_transit_time(plane, plane.current_location, hub_icao)
                
                hub_jobs.append(job)
        
        return hub_jobs
    
    def _get_hub_job_threshold(self, plane_type: str) -> float:
        """Get hub job threshold based on plane type - smaller planes need lower thresholds."""
        thresholds = {
            'Cessna Longitude': 1.0,
            'Cargo 400M': 2.0, 
            'Rockwell B-1 Lancer': 3.0,
            'Concorde': 5.0,
            'VSKYLABS Skyscenders 76': 2.0,
            'Hercules H-4': 1.0,
            'Antonov AN-225-210': 1000.0,  # Keep high for Antonovs
            'Lockheed SR-71': 10.0
        }
        return thresholds.get(plane_type, 5.0)  # Default threshold for unknown types
    
    def _calculate_hub_transit_time(self, plane: PlaneInfo, from_icao: str, to_icao: str) -> float:
        """Calculate transit time from current location to a job hub using actual distance and optimized speed."""
        # Calculate actual distance between airports
        distance = self._calculate_distance_between_airports(from_icao, to_icao)
        
        if distance <= 0:
            # Fallback to a reasonable estimate if distance calculation fails
            return 1.0
        
        # Get optimized speed for transit leg (empty cargo, default airport sizes)
        if self.optimizer and plane.plane_type:
            transit_speed = self.optimizer.get_optimized_speed(
                plane.plane_type, distance, 0.0, 3, 3
            )
            # If optimizer returns 0 (no performance data), fall back to base speed
            if transit_speed <= 0:
                transit_speed = self._get_base_speed(plane.plane_type)
        else:
            transit_speed = self._get_base_speed(plane.plane_type)
        
        # Calculate actual transit time
        if transit_speed > 0:
            return distance / transit_speed
        else:
            # Fallback if no speed available
            return 1.0
    
    def find_multi_job_combinations(self, plane: PlaneInfo, current_location: str, 
                                   remaining_hours: float, plane_spec: Dict[str, Any]) -> List[MultiJobLeg]:
        """Find optimal combinations of jobs that share the same origin and destination."""
        # Get all feasible jobs from anywhere (not just current location)
        available_jobs = [job for job in plane.feasible_jobs 
                         if job.job_id not in self._used_jobs_tracking and job.legs]
        
        if not available_jobs:
            return []
        
        multi_job_legs = []
        
        # Group jobs by origin-destination pairs
        route_groups = defaultdict(list)
        for job in available_jobs:
            # Only consider single-leg jobs for multi-job legs (simpler logic)
            if len(job.legs) == 1:
                leg = job.legs[0]
                route_key = (leg['from_icao'].upper(), leg['to_icao'].upper())
                route_groups[route_key].append(job)
        
        # Create multi-job legs for routes with multiple jobs
        for (from_icao, to_icao), jobs in route_groups.items():
            if len(jobs) > 1:
                # Calculate distance for this route
                distance = jobs[0].legs[0]['distance_nm']
                
                # Find the best combination of jobs for this route
                best_combination = self._find_best_job_combination(jobs, plane_spec, distance, remaining_hours)
                
                if best_combination and len(best_combination) > 1:
                    # Create a multi-job leg for this combination
                    multi_leg = self._create_simple_multi_job_leg(best_combination, from_icao, to_icao, distance)
                    if multi_leg:
                        multi_job_legs.append(multi_leg)
        
        return multi_job_legs
    
    def _find_best_job_combination(self, jobs: List[JobInfo], plane_spec: Dict[str, Any], 
                                  distance: float, remaining_hours: float) -> Optional[List[JobInfo]]:
        """Find the best combination of jobs for the same route that fits within plane limits."""
        if len(jobs) < 2:
            return None
        
        # Sort jobs by efficiency (pay + xp per hour) to try most profitable first
        sorted_jobs = sorted(jobs, key=lambda job: (job.pay + job.xp) / job.flight_hours, reverse=True)
        
        best_combination = []
        best_profit = 0
        total_payload = 0
        total_time = 0
        
        # Try adding jobs one by one, starting with the most efficient
        for job in sorted_jobs:
            # Check if adding this job would exceed limits
            new_total_payload = total_payload + job.total_payload_lbs
            new_total_time = total_time + job.flight_hours
            
            # Check payload capacity for this distance
            max_payload = self._calculate_payload_range_limit(plane_spec, distance)
            if new_total_payload > max_payload:
                continue  # Skip this job - would exceed payload capacity
            
            # Check time limits
            if new_total_time > remaining_hours:
                continue  # Skip this job - would exceed time limit
            
            # Add this job to the combination
            best_combination.append(job)
            total_payload = new_total_payload
            total_time = new_total_time
            best_profit += job.pay + job.xp
        
        # Return the combination if it has multiple jobs and is profitable
        if len(best_combination) > 1 and best_profit > 0:
            return best_combination
        
        return None
    
    def _create_simple_multi_job_leg(self, jobs: List[JobInfo], from_icao: str, to_icao: str, 
                                   distance: float) -> Optional[MultiJobLeg]:
        """Create a multi-job leg from jobs that share the same origin and destination."""
        if len(jobs) < 2:
            return None
        
        # Calculate total payload and flight time
        total_payload = sum(job.total_payload_lbs for job in jobs)
        total_pay = sum(job.pay for job in jobs)
        total_xp = sum(job.xp for job in jobs)
        
        # Use the flight time from the first job (all should be the same for same route)
        flight_hours = jobs[0].flight_hours
        
        # Calculate average speed
        speed_kts = distance / flight_hours if flight_hours > 0 else jobs[0].speed_kts
        
        # Create the multi-job leg
        multi_leg = MultiJobLeg(
            from_icao=from_icao,
            to_icao=to_icao,
            distance_nm=distance,
            flight_hours=flight_hours,
            speed_kts=speed_kts
        )
        
        # Add all jobs to the leg
        for job in jobs:
            multi_leg.add_job(job)
        
        return multi_leg
    
    def _create_chained_multi_job_leg(self, jobs: List[JobInfo]) -> Optional[MultiJobLeg]:
        """Create a multi-job leg from a chained combination of jobs."""
        if len(jobs) < 2:
            return None
        
        # For chained jobs, we need to calculate the combined route and timing
        start_job = jobs[0]
        end_job = jobs[-1]
        
        # Calculate total distance and time
        total_distance = sum(leg['distance_nm'] for job in jobs for leg in job.legs)
        total_hours = sum(job.flight_hours for job in jobs)
        
        # Calculate average speed
        avg_speed = total_distance / total_hours if total_hours > 0 else start_job.speed_kts
        
        # Create the multi-job leg from start to end
        multi_leg = MultiJobLeg(
            from_icao=start_job.departure,
            to_icao=end_job.destination,
            distance_nm=total_distance,
            flight_hours=total_hours,
            speed_kts=avg_speed
        )
        
        # Add all jobs to the leg
        for job in jobs:
            multi_leg.add_job(job)
        
        return multi_leg
    
    def _find_simple_job_combination(self, jobs: List[JobInfo], plane_spec: Dict[str, Any], 
                                   remaining_hours: float) -> Optional[List[JobInfo]]:
        """Find a smart combination of jobs using improved greedy approach."""
        if not jobs:
            return None
        
        # Sort jobs by pure efficiency score (pay+XP per hour)
        def job_efficiency_score(job):
            return (job.pay + job.xp) / job.flight_hours if job.flight_hours > 0 else 0
        
        jobs.sort(key=job_efficiency_score, reverse=True)
        
        combination = []
        total_payload = 0.0
        first_leg = jobs[0].legs[0]
        distance = first_leg['distance_nm']
        max_payload = self._calculate_payload_range_limit(plane_spec, distance)
        
        # Smart greedy: prioritize jobs that maximize efficiency and payload utilization
        for job in jobs:
            # Check if job fits within remaining time
            if job.flight_hours > remaining_hours:
                continue
            
            # Check payload capacity
            new_total_payload = total_payload + job.total_payload_lbs
            if new_total_payload <= max_payload:
                combination.append(job)
                total_payload = new_total_payload
                
                # Try to fill payload efficiently - stop if we're using >80% of capacity
                if total_payload > max_payload * 0.8:
                    break
                
                # Limit to 4 jobs max for performance
                if len(combination) >= 4:
                    break
        
        # Return combination if we have multiple jobs
        if len(combination) >= 2:
            return combination
        
        # Don't return single jobs here - they should be handled as regular jobs
        return None
    
    def _is_job_combination_feasible(self, jobs: List[JobInfo], plane_spec: Dict[str, Any], 
                                   remaining_hours: float) -> bool:
        """Check if a combination of jobs is feasible."""
        if not jobs:
            return False
        
        # All jobs must have the same first leg
        first_leg = jobs[0].legs[0]
        for job in jobs[1:]:
            if (job.legs[0]['from_icao'] != first_leg['from_icao'] or 
                job.legs[0]['to_icao'] != first_leg['to_icao']):
                return False
        
        # Check total payload capacity
        total_payload = sum(job.total_payload_lbs for job in jobs)
        max_payload = self._calculate_payload_range_limit(plane_spec, first_leg['distance_nm'])
        if total_payload > max_payload:
            return False
        
        # Check time constraints
        total_flight_hours = jobs[0].flight_hours  # All jobs share the same flight time
        if total_flight_hours > remaining_hours:
            return False
        
        # Check that all jobs can be delivered before work order completion
        # This is handled by ensuring jobs are sorted by delivery waypoint
        return True
    
    def _get_job_priority_score(self, job: JobInfo) -> float:
        """Get priority score for a job (simplified version)."""
        # Use balance score as default, but this could be customized based on plane priority
        return job.balance_score
    
    def _get_job_chaining_potential(self, destination: str) -> int:
        """Get the number of available jobs from a destination airport."""
        # Direct database query (now using RAM database for speed)
        with db_mod.connect(self.db_path) as conn:
            # Count jobs from this destination that are not used
            if not self._used_jobs_tracking:
                count = conn.execute("""
                    SELECT COUNT(DISTINCT jl.job_id)
                    FROM job_legs jl
                    WHERE jl.from_icao = ?
                """, [destination]).fetchone()[0]
            else:
                count = conn.execute("""
                    SELECT COUNT(DISTINCT jl.job_id)
                    FROM job_legs jl
                    WHERE jl.from_icao = ? 
                    AND jl.job_id NOT IN ({})
                """.format(','.join('?' * len(self._used_jobs_tracking))), 
                [destination] + list(self._used_jobs_tracking)).fetchone()[0]
            return count
    
    def _get_sequential_route_bonus(self, current_location: str, job_departure: str, job_destination: str) -> float:
        """Calculate bonus for maintaining sequential routes."""
        # Perfect sequential route: job starts where we are
        if job_departure.upper() == current_location.upper():
            return 1.5  # 50% bonus for perfect sequential routes
        
        # Partial sequential route: job ends where we are (would need transit to start)
        if job_destination.upper() == current_location.upper():
            return 1.2  # 20% bonus for partial sequential routes
        
        # No sequential connection
        return 1.0  # No bonus
    
    
    def _calculate_payload_range_limit(self, plane_spec: Dict[str, Any], distance_nm: float) -> float:
        """Calculate maximum payload for given distance (copied from score_jobs.py)."""
        max_range = max(plane_spec["range1_nm"] or 0, plane_spec["range2_nm"] or 0)
        max_payload = max(plane_spec["payload1_lbs"] or 0, plane_spec["payload2_lbs"] or 0)
        
        if distance_nm > max_range:
            return 0
        
        if plane_spec["range1_nm"] == plane_spec["range2_nm"]:
            return max_payload
        
        if plane_spec["payload1_lbs"] == plane_spec["payload2_lbs"]:
            return plane_spec["payload1_lbs"]
        
        # Sort the range/payload pairs
        if plane_spec["range1_nm"] > plane_spec["range2_nm"]:
            r1, p1 = plane_spec["range2_nm"], plane_spec["payload2_lbs"]
            r2, p2 = plane_spec["range1_nm"], plane_spec["payload1_lbs"]
        else:
            r1, p1 = plane_spec["range1_nm"], plane_spec["payload1_lbs"]
            r2, p2 = plane_spec["range2_nm"], plane_spec["payload2_lbs"]
        
        if r1 == r2:
            return max_payload if distance_nm <= r1 else 0
        
        if distance_nm <= r1:
            return max(p1, p2)
        elif distance_nm >= r2:
            return min(p1, p2)
        else:
            m = (p2 - p1) / (r2 - r1)
            calculated_payload = p1 + m * (distance_nm - r1)
            return min(calculated_payload, max_payload)

    def generate_work_order(self, plane: PlaneInfo, max_hours: float = 24.0, 
                           epsilon_hours: float = 0.5) -> WorkOrder:
        """Generate a work order for a single plane with multi-job support."""
        work_order = WorkOrder(
            plane_id=plane.plane_id,
            plane_registration=plane.registration,
            plane_type=plane.plane_type,
            priority=plane.priority
        )
        
        current_location = plane.current_location
        remaining_hours = max_hours
        max_transit_legs = 3  # Allow up to 3 transit legs to find better job chains
        
        # Get plane specifications for multi-job validation
        plane_spec = self._get_plane_specs_for_plane(plane)
        
        # First pass: compare multi-job combinations vs single jobs and pick the most profitable
        max_iterations = 12  # Allow more iterations for longer work orders with transit legs
        iteration_count = 0
        
        while remaining_hours > 0.5 and iteration_count < max_iterations:  # Fill time more reasonably
            
            # Find all available options
            multi_job_legs = self.find_multi_job_combinations(plane, current_location, remaining_hours, plane_spec)
            single_jobs = self.find_chainable_jobs(plane, current_location, remaining_hours, epsilon_hours)
            
            # Filter options that fit within remaining time
            feasible_multi_job_legs = [leg for leg in multi_job_legs if leg.flight_hours <= (remaining_hours + epsilon_hours)]
            feasible_single_jobs = [job for job in single_jobs if work_order.can_add_job(job, max_hours, epsilon_hours)]
            
            # Find jobs from current location (most efficient approach)
            jobs_from_current = [job for job in feasible_single_jobs 
                               if job.legs and job.legs[0]['from_icao'].upper() == current_location.upper()]
            
            # Debug output for first few planes and BONE planes
            if iteration_count == 0 and ('ANT' in plane.registration or 'BONE' in plane.registration or 'SKY' in plane.registration):
                print(f"    DEBUG {plane.registration}: {len(multi_job_legs)} multi-job legs, {len(single_jobs)} single jobs available")
                print(f"    DEBUG {plane.registration}: Jobs from current location: {len(jobs_from_current)}")
                
                # Debug multi-job leg creation
                if len(multi_job_legs) == 0 and len(jobs_from_current) > 1:
                    print(f"    DEBUG {plane.registration}: No multi-job legs despite {len(jobs_from_current)} jobs from {current_location}")
                    # Show first few jobs and their legs
                    for i, job in enumerate(jobs_from_current[:3]):
                        if job.legs:
                            leg_info = f"{job.legs[0]['from_icao']} -> {job.legs[0]['to_icao']}"
                            print(f"    DEBUG {plane.registration}: Job {i+1}: {leg_info}, payload: {job.total_payload_lbs:.0f}lbs")
                
                if single_jobs:
                    print(f"    DEBUG {plane.registration}: Best single job score: {work_order.get_priority_score(single_jobs[0]):.2f}")
                if multi_job_legs:
                    best_leg = max(multi_job_legs, key=lambda leg: sum(job.pay + job.xp for job in leg.jobs) / leg.flight_hours)
                    print(f"    DEBUG {plane.registration}: Best multi-job leg: {len(best_leg.jobs)} jobs, {sum(job.pay + job.xp for job in best_leg.jobs) / best_leg.flight_hours:.2f} efficiency")
                    
                    # Additional debug for BONE planes
                    if 'BONE' in plane.registration:
                        print(f"    DEBUG {plane.registration}: Current location: {current_location}")
                        print(f"    DEBUG {plane.registration}: Remaining hours: {remaining_hours:.2f}")
                        if jobs_from_current:
                            first_job = jobs_from_current[0]
                            print(f"    DEBUG {plane.registration}: First job flight hours: {first_job.flight_hours:.2f}")
                            print(f"    DEBUG {plane.registration}: Job fits time: {first_job.flight_hours <= (remaining_hours + 0.5)}")
                    
                    # Additional debug for SKYLABS planes - focus on XP
                    if 'SKY' in plane.registration:
                        print(f"    DEBUG {plane.registration}: SKYLABS - Looking for high-XP jobs")
                        best_leg_xp = sum(job.xp for job in best_leg.jobs)
                        best_leg_pay = sum(job.pay for job in best_leg.jobs)
                        print(f"    DEBUG {plane.registration}: Best multi-job leg XP: {best_leg_xp}, Pay: ${best_leg_pay:,.0f}")
                        
                        # Show top 5 multi-leg jobs by XP
                        if multi_job_legs:
                            print(f"    DEBUG {plane.registration}: Top 5 multi-leg jobs by XP:")
                            sorted_multi = sorted(multi_job_legs, key=lambda x: sum(job.xp for job in x.jobs), reverse=True)[:5]
                            for i, job in enumerate(sorted_multi, 1):
                                job_xp = sum(j.xp for j in job.jobs)
                                job_pay = sum(j.pay for j in job.jobs)
                                print(f"      {i}. {len(job.jobs)} legs, {job_xp} XP, ${job_pay:,.0f} pay, {sum(j.pay + j.xp for j in job.jobs) / job.flight_hours:.2f} efficiency")
                        
                        # Show top 5 single jobs by XP
                        if single_jobs:
                            print(f"    DEBUG {plane.registration}: Top 5 single jobs by XP:")
                            sorted_single = sorted(single_jobs, key=lambda x: x.xp, reverse=True)[:5]
                            for i, job in enumerate(sorted_single, 1):
                                print(f"      {i}. {job.xp} XP, ${job.pay:,.0f} pay, {work_order.get_priority_score(job):.2f} score")
            
            # Sequential route-first selection with efficiency consideration
            best_option = None
            best_efficiency_score = -1
            best_type = None
            
            # 1. Check multi-job legs that start from current location - prioritize sequential routes
            for leg in feasible_multi_job_legs:
                # Only consider multi-job legs that start from current location
                if leg.from_icao.upper() != current_location.upper():
                    continue
                
                # Score based on efficiency (pay+XP per hour)
                total_value = sum(job.pay + job.xp for job in leg.jobs)
                efficiency_score = total_value / leg.flight_hours if leg.flight_hours > 0 else 0
                
                # Bonus for multi-job combinations (they're more efficient)
                if len(leg.jobs) > 1:
                    efficiency_score *= 1.2
                
                # Sequential routes naturally have higher efficiency (no transit time needed)
                # No artificial bonus needed - let natural efficiency determine the best choice
                
                if efficiency_score > best_efficiency_score:
                    best_efficiency_score = efficiency_score
                    best_option = leg
                    best_type = "multi_job"
            
            # 2. Check local single jobs - prioritize sequential routes
            for job in jobs_from_current:
                # Use the plane's priority-based score (pay/XP/balanced per hour)
                efficiency_score = work_order.get_priority_score(job)
                
                # Sequential routes naturally have higher efficiency (no transit time needed)
                # No artificial bonus needed - let natural efficiency determine the best choice
                
                # Additional bonus for jobs that lead to good chaining opportunities
                chaining_potential = self._get_job_chaining_potential(job.destination)
                if chaining_potential > 0:
                    efficiency_score *= (1.0 + min(chaining_potential / 20.0, 0.2))  # Reduced to 20% bonus
                
                if efficiency_score > best_efficiency_score:
                    best_efficiency_score = efficiency_score
                    best_option = job
                    best_type = "single_job"
            
            # Add the best option
            if best_option and best_type == "multi_job":
                # Debug: Show multi-job leg selection
                if 'ANT-07' in plane.registration or 'ANT-08' in plane.registration or 'ANT-06' in plane.registration:
                    print(f"    DEBUG {plane.registration}: Selected multi-job leg from {best_option.from_icao} to {best_option.to_icao} (current: {current_location})")
                
                # Add the multi-job leg
                work_order.add_multi_job_leg(best_option)
                
                # Mark all jobs in this leg as used (in memory only)
                for job in best_option.jobs:
                    self._used_jobs_tracking.add(job.job_id)
                
                # Update location and remaining time
                current_location = best_option.to_icao
                remaining_hours = max_hours - work_order.total_hours
                iteration_count += 1
                continue
            elif best_option and best_type == "single_job":
                # Debug: Show single job selection
                if 'ANT-07' in plane.registration or 'ANT-08' in plane.registration or 'ANT-06' in plane.registration:
                    print(f"    DEBUG {plane.registration}: Selected single job from {best_option.legs[0]['from_icao']} to {best_option.destination} (current: {current_location})")
                
                # Add the single job
                work_order.add_job(best_option, work_order.total_hours)
                self._used_jobs_tracking.add(best_option.job_id)
                current_location = best_option.destination
                remaining_hours = max_hours - work_order.total_hours
                iteration_count += 1
                continue
            
            # 3. If no local options found, consider ALL available jobs anywhere to build the most efficient work order
            transit_legs_used = sum(1 for job in work_order.jobs if job.source == "transit")
            if transit_legs_used < max_transit_legs:
                # Get ALL feasible jobs from anywhere, not just major hubs
                all_available_jobs = [job for job in plane.feasible_jobs if job.job_id not in self._used_jobs_tracking]
                
                # Debug output for planes having issues
                if iteration_count == 0 and ('ANT' in plane.registration or 'BONE' in plane.registration):
                    print(f"    DEBUG {plane.registration}: Available jobs anywhere: {len(all_available_jobs)}")
                    print(f"    DEBUG {plane.registration}: Best local efficiency: {best_efficiency_score:.2f}")
                
                if all_available_jobs:
                    # Find the best job from anywhere that helps build the most efficient work order
                    best_remote_job = None
                    best_remote_score = -1
                    
                    for job in all_available_jobs:
                        # Skip jobs that start from current location (already considered in local jobs)
                        if job.legs and job.legs[0]['from_icao'].upper() == current_location.upper():
                            continue  # Already considered in local jobs
                        
                        # Calculate transit time to this job's starting location
                        transit_time = self._calculate_hub_transit_time(plane, current_location, job.legs[0]['from_icao'])
                        total_time = job.flight_hours + transit_time
                        
                        if total_time <= (remaining_hours + epsilon_hours):
                            # Score based on overall efficiency including transit cost
                            job_efficiency = work_order.get_priority_score(job)
                            overall_efficiency = (job_efficiency * job.flight_hours) / total_time
                            
                            # Natural efficiency limiting - transit time naturally reduces overall efficiency
                            # This allows the algorithm to find truly optimal solutions
                            if overall_efficiency > best_remote_score and overall_efficiency > 0:
                                best_remote_score = overall_efficiency
                                best_remote_job = job
                                best_remote_job.hub_penalty_hours = transit_time
                    
                    # Debug remote job selection
                    if 'BONE' in plane.registration and best_remote_job:
                        transit_efficiency = best_remote_job.flight_hours / (best_remote_job.flight_hours + best_remote_job.hub_penalty_hours)
                        print(f"    DEBUG {plane.registration}: Remote job efficiency: {best_remote_score:.2f}, transit_eff={transit_efficiency:.3f}")
                    
                    if best_remote_job:
                        # Calculate actual distance for transit leg
                        transit_distance = self._calculate_distance_between_airports(
                            current_location, best_remote_job.legs[0]['from_icao']
                        )
                        
                        # Calculate actual flight hours for transit leg using optimized speed if available
                        if self.optimizer and plane.plane_type:
                            # Use optimized speed for transit leg (empty cargo, default airport sizes)
                            transit_speed = self.optimizer.get_optimized_speed(
                                plane.plane_type, transit_distance, 0.0, 3, 3
                            )
                            # If optimizer returns 0 (no performance data), fall back to base speed
                            if transit_speed <= 0:
                                transit_speed = self._get_base_speed(plane.plane_type)
                        else:
                            transit_speed = self._get_base_speed(plane.plane_type)
                        
                        transit_flight_hours = transit_distance / transit_speed if transit_speed > 0 else best_remote_job.hub_penalty_hours
                        
                        # Add transit time as a "virtual job" to account for flying to job location
                        transit_job = JobInfo(
                            job_id=f"transit_{plane.plane_id}_{best_remote_job.legs[0]['from_icao']}",
                            departure=current_location,
                            destination=best_remote_job.legs[0]['from_icao'],
                            pay=0.0,
                            xp=0.0,
                            distance_nm=transit_distance,
                            pay_per_hour=0.0,
                            xp_per_hour=0.0,
                            balance_score=0.0,
                            flight_hours=transit_flight_hours,
                            time_remaining_hours=999.0,  # Transit jobs don't expire
                            source="transit",
                            speed_kts=transit_speed,
                            route=f"{current_location} -> {best_remote_job.legs[0]['from_icao']}",
                            legs_count=0,  # Transit legs don't count as actual job legs
                            legs=[]  # Empty legs for transit jobs
                        )
                        
                        work_order.add_job(transit_job, work_order.total_hours)
                        current_location = best_remote_job.legs[0]['from_icao']
                        remaining_hours = max_hours - work_order.total_hours
                        
                        # Now add the actual remote job
                        # Debug: Show remote job selection
                        if 'ANT-07' in plane.registration or 'ANT-08' in plane.registration or 'ANT-06' in plane.registration:
                            print(f"    DEBUG {plane.registration}: Added remote job from {best_remote_job.legs[0]['from_icao']} to {best_remote_job.destination} with transit leg")
                        
                        work_order.add_job(best_remote_job, work_order.total_hours)
                        self._used_jobs_tracking.add(best_remote_job.job_id)
                        current_location = best_remote_job.destination
                        remaining_hours = max_hours - work_order.total_hours
                        iteration_count += 1
                        continue
            
            # No more options available
            if 'ANT' in plane.registration:
                print(f"    DEBUG {plane.registration}: No more options found. Remaining hours: {remaining_hours:.2f}")
            break
        
        # Second pass: try to fill remaining time with smaller jobs (more aggressive)
        if remaining_hours > 0.2:  # More aggressive about filling remaining time
            # Use the work order's current location, not the loop's current_location variable
            actual_current_location = work_order.get_current_location() if work_order.jobs else plane.current_location
            
            # Debug: Show second pass optimization
            if 'ANT-07' in plane.registration or 'ANT-08' in plane.registration or 'ANT-06' in plane.registration:
                print(f"    DEBUG {plane.registration}: Starting second pass optimization from {actual_current_location} with {remaining_hours:.2f} hours remaining")
            
            self._optimize_work_order_second_pass_simple(work_order, plane, actual_current_location, 
                                                        max_hours, epsilon_hours)
        
        # Remove trailing transit legs (they provide no Pay/XP value)
        self._remove_trailing_transit_legs(work_order)
        
        # Optimize job order to minimize penalties while preserving chaining
        if len(work_order.jobs) > 1:
            # Debug: Show job order before optimization
            if 'ANT-07' in plane.registration or 'ANT-08' in plane.registration or 'ANT-06' in plane.registration:
                print(f"    DEBUG {plane.registration}: Job order before optimization:")
                print(f"      Regular jobs: {len(work_order.jobs)}")
                for i, job in enumerate(work_order.jobs):
                    if hasattr(job, 'legs') and job.legs:
                        from_loc = job.legs[0]['from_icao']
                        to_loc = job.destination
                        print(f"        {i+1}. {from_loc} -> {to_loc} ({job.source})")
                print(f"      Multi-job legs: {len(work_order.multi_job_legs)}")
                for i, leg in enumerate(work_order.multi_job_legs):
                    print(f"        {i+1}. {leg.from_icao} -> {leg.to_icao} (multi-job leg)")
            
            self._optimize_job_order_for_penalties(work_order)
            
            # Debug: Show job order after optimization
            if 'ANT-07' in plane.registration or 'ANT-08' in plane.registration or 'ANT-06' in plane.registration:
                print(f"    DEBUG {plane.registration}: Job order after optimization:")
                for i, job in enumerate(work_order.jobs):
                    if hasattr(job, 'legs') and job.legs:
                        from_loc = job.legs[0]['from_icao']
                        to_loc = job.destination
                        print(f"      {i+1}. {from_loc} -> {to_loc} ({job.source})")
        
        return work_order
    
    def _get_airport_sizes(self, from_icao: str, to_icao: str) -> Tuple[int, int]:
        """Get airport sizes for departure and destination airports with caching."""
        # Check cache first
        if from_icao not in self._airport_size_cache:
            with db_mod.connect(self.db_path) as conn:
                dep_row = conn.execute("SELECT size FROM airports WHERE icao = ?", (from_icao,)).fetchone()
                self._airport_size_cache[from_icao] = int(dep_row[0]) if dep_row and dep_row[0] is not None else 3
        
        if to_icao not in self._airport_size_cache:
            with db_mod.connect(self.db_path) as conn:
                dest_row = conn.execute("SELECT size FROM airports WHERE icao = ?", (to_icao,)).fetchone()
                self._airport_size_cache[to_icao] = int(dest_row[0]) if dest_row and dest_row[0] is not None else 3
        
        return self._airport_size_cache[from_icao], self._airport_size_cache[to_icao]
    
    def _get_plane_specs_for_plane(self, plane: PlaneInfo) -> Dict[str, Any]:
        """Get plane specifications for a plane."""
        # Direct database query (now using RAM database for speed)
        with db_mod.connect(self.db_path) as conn:
            # Get plane specs from the database
            row = conn.execute("SELECT * FROM plane_specs WHERE plane_type = ?", (plane.plane_type,)).fetchone()
            if row:
                columns = [d[1] for d in conn.execute("PRAGMA table_info(plane_specs)").fetchall()]
                spec_dict = dict(zip(columns, row))
                return {
                    "speed_kts": spec_dict.get("cruise_speed_kts") or 0,
                    "min_airport_size": spec_dict.get("min_airport_size") or 0,
                    "range1_nm": spec_dict.get("range1_nm") or 0,
                    "payload1_lbs": spec_dict.get("payload1_lbs") or 0,
                    "range2_nm": spec_dict.get("range2_nm") or 0,
                    "payload2_lbs": spec_dict.get("payload2_lbs") or 0,
                    "priority": spec_dict.get("priority") or "balance"
                }
            else:
                # Fallback to default values
                return {
                    "speed_kts": 200.0,
                    "min_airport_size": 0,
                    "range1_nm": 1000.0,
                    "payload1_lbs": 1000.0,
                    "range2_nm": 1000.0,
                    "payload2_lbs": 1000.0,
                    "priority": "balance"
                }
    
    def _optimize_work_order_second_pass_simple(self, work_order: WorkOrder, plane: PlaneInfo,
                                              current_location: str, max_hours: float, 
                                              epsilon_hours: float) -> None:
        """Improved second pass optimization to better fill work orders."""
        remaining_hours = max_hours - work_order.total_hours
        
        # Try to add jobs if there's time remaining and good efficiency options
        max_attempts = 3 if remaining_hours > 4.0 else 1  # Fewer attempts, focus on efficiency
        
        for attempt in range(max_attempts):
            if remaining_hours < 0.5:  # Stop if less than 30 minutes left (reasonable threshold)
                break
                
            chainable_jobs = self.find_chainable_jobs(plane, current_location, remaining_hours, epsilon_hours)
            
            if not chainable_jobs:
                break
            
            # Find the best job that fits in remaining time
            best_job = None
            best_score = -1
            
            for job in chainable_jobs:
                if job.flight_hours <= (remaining_hours + epsilon_hours):
                    # Score based purely on efficiency (plane's priority)
                    score = work_order.get_priority_score(job)
                    
                    if score > best_score:
                        best_score = score
                        best_job = job
            
            if best_job:
                # Debug: Show second pass job selection
                if 'ANT-07' in plane.registration or 'ANT-08' in plane.registration or 'ANT-06' in plane.registration:
                    print(f"    DEBUG {plane.registration}: Second pass added job from {best_job.legs[0]['from_icao']} to {best_job.destination} (current: {current_location})")
                
                # Add the job
                work_order.add_job(best_job, work_order.total_hours)
                self._used_jobs_tracking.add(best_job.job_id)
                
                # Update for next iteration
                current_location = best_job.destination
                remaining_hours = max_hours - work_order.total_hours
            else:
                break
    
    def _remove_trailing_transit_legs(self, work_order: WorkOrder) -> None:
        """Remove trailing transit legs from work order since they provide no Pay/XP value."""
        if not work_order.jobs:
            return
        
        # Debug: Show jobs before removing trailing transit legs
        if 'ANT-07' in str(work_order.jobs[0].job_id if work_order.jobs else ''):
            print(f"    DEBUG: Before removing trailing transit legs: {len(work_order.jobs)} jobs")
            for i, job in enumerate(work_order.jobs):
                if hasattr(job, 'legs') and job.legs:
                    from_loc = job.legs[0]['from_icao']
                    to_loc = job.destination
                    print(f"      {i+1}. {from_loc} -> {to_loc} ({job.source})")
        
        # Remove transit legs from the end of the work order
        removed_count = 0
        while work_order.jobs and work_order.jobs[-1].source == "transit":
            removed_job = work_order.jobs.pop()
            removed_count += 1
            # Update totals
            work_order.total_hours -= removed_job.flight_hours
            work_order.total_pay -= removed_job.pay
            work_order.total_xp -= removed_job.xp
            work_order.total_adjusted_pay -= removed_job.adjusted_pay_total
        
        # Debug: Show jobs after removing trailing transit legs
        if 'ANT-07' in str(work_order.jobs[0].job_id if work_order.jobs else '') and removed_count > 0:
            print(f"    DEBUG: Removed {removed_count} trailing transit legs: {len(work_order.jobs)} jobs remaining")
            for i, job in enumerate(work_order.jobs):
                if hasattr(job, 'legs') and job.legs:
                    from_loc = job.legs[0]['from_icao']
                    to_loc = job.destination
                    print(f"      {i+1}. {from_loc} -> {to_loc} ({job.source})")
        
        # Recalculate totals to ensure consistency
        self._recalculate_work_order_totals(work_order)
    
    def _optimize_work_order_second_pass(self, work_order: WorkOrder, plane: PlaneInfo,
                                        current_location: str, max_hours: float, 
                                        epsilon_hours: float) -> None:
        """Second pass optimization to fill remaining time."""
        remaining_hours = max_hours - work_order.total_hours
        
        while remaining_hours > 0.1:
            chainable_jobs = self.find_chainable_jobs(plane, current_location, remaining_hours, epsilon_hours)
            
            if not chainable_jobs:
                break
            
            # Find the best job that fits in remaining time
            best_job = None
            best_score = -1
            
            for job in chainable_jobs:
                if job.flight_hours <= (remaining_hours + epsilon_hours):
                    score = work_order.get_priority_score(job)
                    if score > best_score:
                        best_score = score
                        best_job = job
            
            if not best_job:
                break
            
            # Add the job
            work_order.add_job(best_job, work_order.total_hours)
            self._used_jobs_tracking.add(best_job.job_id)
            
            # Update for next iteration
            current_location = best_job.destination
            remaining_hours = max_hours - work_order.total_hours
    
    def _optimize_job_order_for_penalties(self, work_order: WorkOrder) -> None:
        """Optimize job order to minimize penalties while preserving chaining.
        
        Only sorts jobs that have the same departure and destination to avoid
        breaking the sequential chaining requirement.
        """
        if len(work_order.jobs) <= 1:
            return
        
        # Group consecutive jobs by departure/destination
        groups = []
        current_group = [work_order.jobs[0]]
        
        # Debug: Show job grouping process
        print(f"    DEBUG: Starting job grouping for {len(work_order.jobs)} jobs")
        for i, job in enumerate(work_order.jobs):
            # Use consistent departure/destination logic
            from_loc = job.legs[0]['from_icao'] if job.legs else job.departure
            to_loc = job.legs[-1]['to_icao'] if job.legs else job.destination
            print(f"      Job {i+1}: {from_loc} -> {to_loc} ({job.source}) [time_remaining: {job.time_remaining_hours:.1f}h]")
        
        for i in range(1, len(work_order.jobs)):
            prev_job = work_order.jobs[i-1]
            curr_job = work_order.jobs[i]
            
            # Use consistent departure/destination logic for comparison
            prev_from = prev_job.legs[0]['from_icao'] if prev_job.legs else prev_job.departure
            prev_to = prev_job.legs[-1]['to_icao'] if prev_job.legs else prev_job.destination
            curr_from = curr_job.legs[0]['from_icao'] if curr_job.legs else curr_job.departure
            curr_to = curr_job.legs[-1]['to_icao'] if curr_job.legs else curr_job.destination
            
            # Check if jobs have same departure and destination
            if prev_from == curr_from and prev_to == curr_to:
                current_group.append(curr_job)
                print(f"    DEBUG: Added job {i+1} to group (same route: {curr_from} -> {curr_to})")
            else:
                # End current group and start new one
                if len(current_group) > 1:
                    groups.append(current_group)
                    print(f"    DEBUG: Closed group with {len(current_group)} jobs")
                current_group = [curr_job]
                print(f"    DEBUG: Started new group with job {i+1} (route: {curr_from} -> {curr_to})")
        
        # Add the last group if it has multiple jobs
        if len(current_group) > 1:
            groups.append(current_group)
        
        # Debug: Show groups found
        print(f"    DEBUG: Found {len(groups)} groups with multiple jobs")
        
        # Sort each group by time remaining (ascending = most urgent first)
        for i, group in enumerate(groups):
            print(f"    DEBUG: Sorting group {i+1} with {len(group)} jobs by time remaining")
            original_order = [f"{j.job_id[:8]}...({j.time_remaining_hours:.1f}h)" for j in group]
            group.sort(key=lambda j: j.time_remaining_hours)
            sorted_order = [f"{j.job_id[:8]}...({j.time_remaining_hours:.1f}h)" for j in group]
            print(f"      Original: {original_order}")
            print(f"      Sorted:   {sorted_order}")
        
        # Rebuild the job list maintaining the overall order
        new_jobs = []
        
        # Debug: Show what groups were found
        print(f"    DEBUG: Found {len(groups)} groups to sort")
        for i, group in enumerate(groups):
            print(f"      Group {i}: {len(group)} jobs")
        
        # Simple approach: iterate through original jobs and replace with sorted versions if they're in groups
        for i, job in enumerate(work_order.jobs):
            job_replaced = False
            
            # Check if this job is in any group
            for group in groups:
                if job in group:
                    # Find the sorted version of this job
                    for sorted_job in group:
                        if sorted_job.job_id == job.job_id:
                            new_jobs.append(sorted_job)
                            job_replaced = True
                            print(f"    DEBUG: Replaced job {i+1} with sorted version")
                            break
                    break
            
            if not job_replaced:
                # Job not in any group, keep original order
                new_jobs.append(job)
                print(f"    DEBUG: Added job {i+1} as-is (not in group)")
        
        # Debug: Show final job count
        print(f"    DEBUG: Rebuilt job list: {len(work_order.jobs)} -> {len(new_jobs)} jobs")
        
        work_order.jobs = new_jobs
        
        # Recalculate totals and penalties with correct accumulated times after reordering
        self._recalculate_work_order_totals(work_order)
    
    def _recalculate_work_order_totals(self, work_order: WorkOrder) -> None:
        """Recalculate work order totals after job reordering."""
        work_order.total_hours = 0.0
        work_order.total_pay = 0.0
        work_order.total_xp = 0.0
        work_order.total_adjusted_pay = 0.0
        
        accumulated_hours = 0.0
        for job in work_order.jobs:
            accumulated_hours += job.flight_hours
            
            # Recalculate penalty based on new position
            job.penalty_amount = work_order._calculate_penalty(job, accumulated_hours)
            job.adjusted_pay_total = job.pay - job.penalty_amount
            job.adjusted_pay_per_hour = job.adjusted_pay_total / job.flight_hours if job.flight_hours > 0 else 0.0
            
            work_order.total_hours += job.flight_hours
            work_order.total_pay += job.pay
            work_order.total_xp += job.xp
            work_order.total_adjusted_pay += job.adjusted_pay_total
        
        # Also include multi-job legs in totals
        for leg in work_order.multi_job_legs:
            work_order.total_hours += leg.flight_hours
            for job in leg.jobs:
                work_order.total_pay += job.pay
                work_order.total_xp += job.xp
                work_order.total_adjusted_pay += job.adjusted_pay_total
    
    def generate_all_work_orders(self, max_hours: float = 24.0, 
                                epsilon_hours: float = 0.5) -> List[WorkOrder]:
        """Generate work orders for all idle planes."""
        idle_planes = self.get_idle_planes()
        work_orders = []
        
        print(f"Found {len(idle_planes)} idle planes")
        
        # Sort planes by efficiency within priority groups: Pay → XP → Balanced
        # This ensures the most efficient planes get first access to the best jobs
        def plane_efficiency_key(plane):
            # Priority order: Pay planes first, then XP, then Balanced
            priority_order = {'pay': 0, 'xp': 1, 'balance': 2}
            priority_score = priority_order.get(plane.priority, 3)
            
            # Within each priority group, sort by plane efficiency
            # Higher efficiency planes (better pay/XP per hour) should go first
            efficiency_scores = {
                'Antonov AN-225-210': 1000,  # Highest pay efficiency
                'Lockheed SR-71': 900,       # High XP efficiency
                'Concorde': 800,             # Good XP efficiency
                'Rockwell B-1 Lancer': 700,  # Good XP efficiency
                'Cargo 400M': 600,           # Moderate efficiency
                'Cessna Longitude': 500,     # Lower efficiency
                'Hercules H-4': 400,         # Lower efficiency
                'VSKYLABS Skyscenders 76': 300  # Lower efficiency
            }
            
            # Sort by priority first, then by efficiency (descending)
            return (priority_score, -efficiency_scores.get(plane.plane_type, 100))
        
        idle_planes.sort(key=plane_efficiency_key)
        
        for i, plane in enumerate(idle_planes, 1):
            print(f"Processing plane {i}/{len(idle_planes)}: {plane.registration} ({plane.plane_type}, priority: {plane.priority})")
            
            # Load feasible jobs for this plane
            self.load_feasible_jobs_for_plane(plane)
            
            # Debug for Concorde and Lancer planes
            if plane.plane_type in ['Concorde', 'Rockwell B-1 Lancer']:
                print(f"  DEBUG {plane.registration}: Total feasible jobs: {len(plane.feasible_jobs)}")
                if plane.feasible_jobs:
                    # Show first few jobs and their locations
                    for j, job in enumerate(plane.feasible_jobs[:3]):
                        if job.legs:
                            leg_info = f"{job.legs[0]['from_icao']} -> {job.destination}"
                            print(f"  DEBUG {plane.registration}: Job {j+1}: {leg_info}, pay: ${job.pay:,.0f}, hours: {job.flight_hours:.1f}h")
                else:
                    print(f"  DEBUG {plane.registration}: No feasible jobs found - checking why...")
                    # Check if there are ANY jobs in the database for this plane type
                    with db_mod.connect(self.db_path) as conn:
                        total_jobs = conn.execute("""
                            SELECT COUNT(*) FROM job_scores js 
                            JOIN jobs j ON j.id = js.job_id 
                            WHERE js.plane_id = ?
                        """, (plane.plane_id,)).fetchone()
                        feasible_jobs = conn.execute("""
                            SELECT COUNT(*) FROM job_scores js 
                            JOIN jobs j ON j.id = js.job_id 
                            WHERE js.plane_id = ? AND js.feasible = 1
                        """, (plane.plane_id,)).fetchone()
                        print(f"  DEBUG {plane.registration}: Total jobs scored: {total_jobs[0]}, Feasible: {feasible_jobs[0]}")
            
            if not plane.feasible_jobs:
                print(f"  No feasible jobs found for {plane.registration}")
                continue
            
            print(f"  Found {len(plane.feasible_jobs)} feasible jobs")
            
            # Generate work order
            work_order = self.generate_work_order(plane, max_hours, epsilon_hours)
            
            # Check if work order has any profitable jobs (regular jobs or multi-job legs)
            total_jobs = len(work_order.jobs) + sum(len(leg.jobs) for leg in work_order.multi_job_legs)
            
            # Filter out work orders that only contain transit jobs or end with transit jobs
            has_profitable_jobs = False
            if work_order.jobs:
                # Check if any regular job has pay or XP
                has_profitable_jobs = any(job.pay > 0 or job.xp > 0 for job in work_order.jobs)
            if work_order.multi_job_legs:
                # Check if any multi-job leg has pay or XP
                has_profitable_jobs = has_profitable_jobs or any(
                    any(job.pay > 0 or job.xp > 0 for job in leg.jobs) 
                    for leg in work_order.multi_job_legs
                )
            
            # Also check that the work order doesn't end with a transit job
            ends_with_transit = False
            if work_order.jobs and work_order.jobs[-1].source == "transit":
                ends_with_transit = True
            
            if total_jobs > 0 and has_profitable_jobs and not ends_with_transit:
                work_orders.append(work_order)
                print(f"  Generated work order with {total_jobs} jobs ({len(work_order.jobs)} regular, {len(work_order.multi_job_legs)} multi-job legs), {work_order.total_hours:.2f} hours")
            else:
                if total_jobs == 0:
                    print(f"  No suitable job chain found for {plane.registration}")
                elif not has_profitable_jobs:
                    print(f"  No profitable jobs found for {plane.registration} (only transit jobs)")
                elif ends_with_transit:
                    print(f"  Work order ends with transit job for {plane.registration} (rejected)")
                # Debug: Check what happened for PANO and BONE planes
                if 'PANO' in plane.registration or 'BONE' in plane.registration:
                    print(f"    DEBUG: {plane.registration} - Feasible jobs: {len(plane.feasible_jobs)}")
                    print(f"    DEBUG: {plane.registration} - Used jobs count: {len(self._used_jobs_tracking)}")
                    if plane.feasible_jobs:
                        available_jobs = [job for job in plane.feasible_jobs if job.job_id not in self._used_jobs_tracking]
                        print(f"    DEBUG: {plane.registration} - Available jobs: {len(available_jobs)}")
                        if available_jobs:
                            print(f"    DEBUG: {plane.registration} - First available job: {available_jobs[0].job_id}")
                            
                        # Debug work order details
                        print(f"    DEBUG: {plane.registration} - Work order jobs: {len(work_order.jobs)}")
                        print(f"    DEBUG: {plane.registration} - Work order multi-job legs: {len(work_order.multi_job_legs)}")
                        if work_order.jobs:
                            print(f"    DEBUG: {plane.registration} - Work order total hours: {work_order.total_hours:.2f}")
                            print(f"    DEBUG: {plane.registration} - Has profitable jobs: {has_profitable_jobs}")
                            print(f"    DEBUG: {plane.registration} - Ends with transit: {ends_with_transit}")
                            if work_order.jobs:
                                last_job = work_order.jobs[-1]
                                print(f"    DEBUG: {plane.registration} - Last job source: {last_job.source}, pay: {last_job.pay}, xp: {last_job.xp}")
                        else:
                            print(f"    DEBUG: {plane.registration} - All jobs already used")
                    else:
                        print(f"    DEBUG: {plane.registration} - No feasible jobs")
        
        # Write used jobs to database only once at the end
        self.write_used_jobs_to_database()
        
        return work_orders


def print_work_orders(work_orders: List[WorkOrder]) -> None:
    """Print work orders in a readable format."""
    if not work_orders:
        print("No work orders generated.")
        return
    
    print(f"\n{'='*80}")
    print(f"GENERATED {len(work_orders)} WORK ORDERS")
    print(f"{'='*80}")
    
    for i, wo in enumerate(work_orders, 1):
        print(f"\nWork Order #{i}")
        print(f"Plane: {wo.plane_registration} ({wo.plane_type})")
        print(f"Priority: {wo.priority.upper()}")
        print(f"Total Time: {wo.total_hours:.2f} hours")
        print(f"Total Pay: ${wo.total_pay:,.0f} (Adjusted: ${wo.total_adjusted_pay:,.0f})")
        print(f"Total XP: {wo.total_xp:,.0f}")
        print(f"Jobs: {len(wo.jobs)}")
        
        if wo.jobs or wo.multi_job_legs:
            print("\nJob Sequence:")
            accumulated_hours = 0.0
            job_counter = 1
            
            # Print regular jobs first
            for job in wo.jobs:
                accumulated_hours += job.flight_hours
                late_indicator = f" [LATE {accumulated_hours - job.time_remaining_hours:.1f}h]" if accumulated_hours > job.time_remaining_hours else ""
                penalty_str = f" penalty=${job.penalty_amount:.0f}" if job.penalty_amount > 0 else ""
                
                print(f"  {job_counter}. {job.departure} -> {job.destination}")
                print(f"     Job: {job.job_id}")
                print(f"     Distance: {job.distance_nm:.0f}nm, Time: {job.flight_hours:.2f}h")
                print(f"     Pay: ${job.pay:,.0f}/hr -> ${job.pay_per_hour:.0f}, XP: {job.xp_per_hour:.0f}/hr")
                print(f"     Time Remaining: {job.time_remaining_hours:.1f}h{late_indicator}{penalty_str}")
                job_counter += 1
            
            # Print multi-job legs
            for leg in wo.multi_job_legs:
                accumulated_hours += leg.flight_hours
                
                # Calculate combined metrics for the leg
                total_pay = sum(job.pay for job in leg.jobs)
                total_xp = sum(job.xp for job in leg.jobs)
                combined_pay_per_hour = total_pay / leg.flight_hours if leg.flight_hours > 0 else 0
                combined_xp_per_hour = total_xp / leg.flight_hours if leg.flight_hours > 0 else 0
                
                print(f"  {job_counter}. {leg.from_icao} -> {leg.to_icao} (MULTI-JOB LEG)")
                print(f"     Distance: {leg.distance_nm:.0f}nm, Time: {leg.flight_hours:.2f}h, Speed: {leg.speed_kts:.0f}kts")
                print(f"     Total Payload: {leg.total_payload_lbs:.0f}lbs")
                print(f"     Combined: ${total_pay:,.0f} total pay, {total_xp:,.0f} total XP")
                print(f"     Combined: ${combined_pay_per_hour:.0f}/hr, {combined_xp_per_hour:.0f} XP/hr")
                print(f"     Jobs ({len(leg.jobs)}):")
                
                for i, job in enumerate(leg.jobs, 1):
                    late_indicator = f" [LATE {accumulated_hours - job.time_remaining_hours:.1f}h]" if accumulated_hours > job.time_remaining_hours else ""
                    penalty_str = f" penalty=${job.penalty_amount:.0f}" if job.penalty_amount > 0 else ""
                    
                    print(f"       {i}. {job.job_id} - Pay: ${job.pay:,.0f}, XP: {job.xp:,.0f}")
                    print(f"          Payload: {job.total_payload_lbs:.0f}lbs, Time Remaining: {job.time_remaining_hours:.1f}h{late_indicator}{penalty_str}")
                
                job_counter += 1
        
        print(f"\n{'-'*60}")


# ----------------------------
# Excel Export Functions
# ----------------------------
def create_excel_workbook():
    """Create a new Excel workbook with headers and formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Work Orders"
        
        # Add headers
        headers = [
            'work_order_id', 'plane_id', 'plane_registration', 'plane_type', 'priority',
            'job_sequence', 'job_id', 'source', 'departure', 'destination', 'distance_nm',
            'flight_hours', 'pay', 'xp', 'pay_per_hour', 'xp_per_hour', 'balance_score',
            'time_remaining_hours', 'penalty_amount', 'adjusted_pay_per_hour', 'adjusted_pay_total',
            'speed_kts', 'min_airport_size', 'route', 'legs_count',
            'accumulated_hours', 'is_late', 'job_type', 'payload_lbs'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        return wb, ws
    except ImportError:
        print("Warning: openpyxl not available. Install with: pip install openpyxl")
        return None, None


def format_excel_workbook(wb, ws):
    """Format Excel workbook with auto-sized columns and styling."""
    try:
        from openpyxl.utils import get_column_letter
        
        # Auto-size columns based on content
        for col in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in range(1, ws.max_row + 1):
                cell_value = str(ws.cell(row=row, column=col).value or "")
                max_length = max(max_length, len(cell_value))
            
            # Set reasonable column widths
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format specific columns for better readability
        # Pay columns (13, 15, 19, 20) - currency format
        for col in [13, 15, 19, 20]:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.number_format = '#,##0.00'
        
        # XP column (14) - 1 decimal place
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=14)
            if cell.value is not None:
                cell.number_format = '0.0'
        
        # Distance column (11) - 1 decimal place
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=11)
            if cell.value is not None:
                cell.number_format = '0.0'
        
        # Flight hours column (12) - 2 decimal places
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=12)
            if cell.value is not None:
                cell.number_format = '0.00'
        
        # Time remaining column (18) - 1 decimal place
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=18)
            if cell.value is not None:
                cell.number_format = '0.0'
        
        return True
    except Exception as e:
        print(f"Warning: Could not format Excel file: {e}")
        return False


def export_to_excel(work_orders: List[WorkOrder], excel_path: str) -> bool:
    """Export work orders to Excel with formatting. Returns True if successful."""
    wb, ws = create_excel_workbook()
    if not wb or not ws:
        return False
    
    try:
        # Add data rows
        for wo_id, wo in enumerate(work_orders, 1):
            accumulated_hours = 0.0
            job_seq = 1
            
            # Export jobs and multi-job legs in execution order
            for item_type, item in wo.execution_order:
                if item_type == "job":
                    # Export regular job
                    job = item
                    accumulated_hours += job.flight_hours
                    is_late = accumulated_hours > job.time_remaining_hours
                    
                    row_data = [
                        wo_id, wo.plane_id, wo.plane_registration, wo.plane_type, wo.priority,
                        job_seq, job.job_id, job.source, job.departure, job.destination, job.distance_nm,
                        job.flight_hours, job.pay, job.xp, job.pay_per_hour, job.xp_per_hour, job.balance_score,
                        job.time_remaining_hours, job.penalty_amount, job.adjusted_pay_per_hour, job.adjusted_pay_total,
                        job.speed_kts, job.min_airport_size or "", job.route, job.legs_count,
                        accumulated_hours, is_late, "single", job.total_payload_lbs
                    ]
                    
                    ws.append(row_data)
                    job_seq += 1
                    
                elif item_type == "multi_leg":
                    # Export multi-job leg
                    leg = item
                    accumulated_hours += leg.flight_hours
                    
                    for i, job in enumerate(leg.jobs, 1):
                        is_late = accumulated_hours > job.time_remaining_hours
                        
                        row_data = [
                            wo_id, wo.plane_id, wo.plane_registration, wo.plane_type, wo.priority,
                            job_seq, job.job_id, job.source, leg.from_icao, leg.to_icao, leg.distance_nm,
                            leg.flight_hours, job.pay, job.xp, job.pay_per_hour, job.xp_per_hour, job.balance_score,
                            job.time_remaining_hours, job.penalty_amount, job.adjusted_pay_per_hour, job.adjusted_pay_total,
                            leg.speed_kts, job.min_airport_size or "", f"{leg.from_icao} -> {leg.to_icao}", job.legs_count,
                            accumulated_hours, is_late, f"multi_{i}/{len(leg.jobs)}", job.total_payload_lbs
                        ]
                        
                        ws.append(row_data)
                    job_seq += 1
        
        # Format the workbook
        format_excel_workbook(wb, ws)
        
        # Save with safe file operations
        def save_operation():
            wb.save(excel_path)
            return True
        
        return safe_file_operation(save_operation, excel_path, "save Excel file")
        
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        return False


def main():
    """Main function."""
    parser = argparse.ArgumentParser(description="Generate work orders for idle planes")
    parser.add_argument("--max-hours", type=float, default=24.0, 
                       help="Maximum hours for work orders (default: 24.0)")
    parser.add_argument("--epsilon-hours", type=float, default=0.5,
                       help="Epsilon soft limit in hours (default: 0.5)")
    parser.add_argument("--output", default="workorders.xlsx", 
                       help="Output file path (default: workorders.xlsx)")
    parser.add_argument("--format", choices=["excel", "csv"], default="excel",
                       help="Output format: excel or csv (default: excel)")
    parser.add_argument("--no-export", action="store_true",
                       help="Skip file export")
    
    args = parser.parse_args()
    
    # Load configuration
    config = load_config()
    
    # Initialize performance optimizer
    optimizer = None
    try:
        import os
        optimizer = PerformanceOptimizer(config.db_path)
        excel_path = os.path.join(os.path.dirname(__file__), 'planes.xlsx')
        if os.path.exists(excel_path):
            print("Loading performance optimization data...")
            optimizer.load_and_process(excel_path)
        else:
            print("No performance optimization data found, using default calculations")
            optimizer = None
    except Exception as e:
        print(f"Warning: Could not initialize performance optimizer: {e}")
        optimizer = None
    
    # Generate work orders
    generator = WorkOrderGenerator(config.db_path, optimizer)
    work_orders = generator.generate_all_work_orders(args.max_hours, args.epsilon_hours)
    
    # Print results
    print_work_orders(work_orders)
    
    # Export to file by default (unless --no-export is specified)
    if not args.no_export and work_orders:
        # Ensure output file has correct extension based on format
        output_file = args.output
        if args.format == "excel" and not output_file.endswith('.xlsx'):
            output_file = output_file.rsplit('.', 1)[0] + '.xlsx'
        elif args.format == "csv" and not output_file.endswith('.csv'):
            output_file = output_file.rsplit('.', 1)[0] + '.csv'
        
        print(f"\n📊 Exporting work orders to {args.format.upper()} format...")
        
        # Remove existing file if it exists
        if not safe_remove_file(output_file):
            print("❌ Cannot proceed without removing the existing file.")
            return 1
        
        # Export based on format
        success = False
        if args.format == "excel":
            success = export_to_excel(work_orders, output_file)
        else:  # csv
            success = export_to_csv(work_orders, output_file)
        
        if success:
            print(f"✅ Work orders exported successfully to: {output_file}")
            
            # Auto-open the file
            print("📂 Opening file...")
            if open_file(output_file):
                print("✅ File opened successfully!")
            else:
                print("ℹ️  File saved but could not be auto-opened.")
        else:
            print("❌ Failed to export work orders.")
            return 1
    elif not work_orders:
        print("\nNo work orders generated.")
    
    return 0


def export_to_csv(work_orders: List[WorkOrder], csv_path: str) -> bool:
    """Export work orders to CSV with safe file operations. Returns True if successful."""
    import csv
    
    def write_csv_operation():
        # Prepare all data in memory first to minimize disk writes
        all_rows = []
        
        # Header
        header = [
            'work_order_id', 'plane_id', 'plane_registration', 'plane_type', 'priority',
            'job_sequence', 'job_id', 'source', 'departure', 'destination', 'distance_nm',
            'flight_hours', 'pay', 'xp', 'pay_per_hour', 'xp_per_hour', 'balance_score',
            'time_remaining_hours', 'penalty_amount', 'adjusted_pay_per_hour', 'adjusted_pay_total',
            'speed_kts', 'min_airport_size', 'route', 'legs_count',
            'accumulated_hours', 'is_late', 'job_type', 'payload_lbs'
        ]
        all_rows.append(header)
        
        # Data
        for wo_id, wo in enumerate(work_orders, 1):
            accumulated_hours = 0.0
            job_seq = 1
            
            # Export jobs and multi-job legs in execution order
            for item_type, item in wo.execution_order:
                if item_type == "job":
                    # Export regular job
                    job = item
                    accumulated_hours += job.flight_hours
                    is_late = accumulated_hours > job.time_remaining_hours
                    
                    all_rows.append([
                        wo_id, wo.plane_id, wo.plane_registration, wo.plane_type, wo.priority,
                        job_seq, job.job_id, job.source, job.departure, job.destination, job.distance_nm,
                        job.flight_hours, job.pay, job.xp, job.pay_per_hour, job.xp_per_hour, job.balance_score,
                        job.time_remaining_hours, job.penalty_amount, job.adjusted_pay_per_hour, job.adjusted_pay_total,
                        job.speed_kts, job.min_airport_size or "", job.route, job.legs_count,
                        accumulated_hours, is_late, "single", job.total_payload_lbs
                    ])
                    job_seq += 1
                    
                elif item_type == "multi_leg":
                    # Export multi-job leg
                    leg = item
                    accumulated_hours += leg.flight_hours
                    
                    for i, job in enumerate(leg.jobs, 1):
                        is_late = accumulated_hours > job.time_remaining_hours
                        
                        all_rows.append([
                            wo_id, wo.plane_id, wo.plane_registration, wo.plane_type, wo.priority,
                            job_seq, job.job_id, job.source, leg.from_icao, leg.to_icao, leg.distance_nm,
                            leg.flight_hours, job.pay, job.xp, job.pay_per_hour, job.xp_per_hour, job.balance_score,
                            job.time_remaining_hours, job.penalty_amount, job.adjusted_pay_per_hour, job.adjusted_pay_total,
                            leg.speed_kts, job.min_airport_size or "", f"{leg.from_icao} -> {leg.to_icao}", job.legs_count,
                            accumulated_hours, is_late, f"multi_{i}/{len(leg.jobs)}", job.total_payload_lbs
                        ])
                    job_seq += 1
        
        # Write all data in a single operation
        with open(csv_path, 'w', newline='', encoding='utf-8', buffering=8192) as f:
            writer = csv.writer(f)
            writer.writerows(all_rows)
        return True
    
    return safe_file_operation(write_csv_operation, csv_path, "write CSV file")


if __name__ == "__main__":
    sys.exit(main())
